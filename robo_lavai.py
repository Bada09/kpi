"""
robo_lavai.py — Robô de Extração de Relatórios LavAI / VendTEF
Executa a cada 6 horas, faz login no portal e salva os dados.

Instalar dependências:
    pip install playwright beautifulsoup4 schedule
    playwright install chromium

Uso:
    python robo_lavai.py              # roda em loop contínuo (a cada 6h)
    python robo_lavai.py --agora      # executa uma vez imediatamente e sai
"""

import asyncio
import json
import re
import sys
import logging
import csv
import urllib.request
import zipfile
import openpyxl
from datetime import datetime, timezone, timedelta
from pathlib import Path

import schedule
import time
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─── Configurações ────────────────────────────────────────────────────────────
LOGIN = "lavai.alugueis"
SENHA = "L@v#35554"

URL_ERP = "https://www.erpvending.com.br/"
URL_DOWNLOAD  = "https://www.portalvendtef.com.br/relatoriogeral/relatorioVendasGeralDownload"
URL_PAYBLU    = "https://www.portalpayblu.com.br/private-label/relatorio-vendas-private-label-download"

# URL do Google Apps Script
APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbwzGx-piTEHw8GdINEU4fFSspAWQU5kh83OrABUsGDBZrd58mOnalEOQxIQNHhs_5GL/exec"

# Configurações API VMPay Cashless
VMPAY_API_URL = "https://vmpay.vertitecnologia.com.br/api/v1/cashless_transactions"
VMPAY_TOKEN   = "ZCT1YHTYDFnscjy5Kt9AdhmJONuSo7KW7oYe4WcF"

SAIDA_JSON        = Path(__file__).parent / "dados_relatorios.json"
LOG_FILE          = Path(__file__).parent / "robo_lavai.log"

# CSVs locais gerados pelo robô (não vão para o Sheets)
CSV_VMPAY_LOCAL    = Path(__file__).parent / "vmpay_local.js"
CSV_VENDTEF_LOCAL  = Path(__file__).parent / "vendtef_local.js"
CSV_PAYBLU_LOCAL   = Path(__file__).parent / "payblu_local.js"
CSV_SQI_LOCAL      = Path(__file__).parent / "sqi_local.js"

TIMEOUT_MS = 30_000   # 30 segundos
# ──────────────────────────────────────────────────────────────────────────────
FUSO_SP = timezone(timedelta(hours=-3))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logging.Formatter.converter = lambda *args: datetime.now(FUSO_SP).timetuple()
log = logging.getLogger(__name__)


# ─── Helpers ──────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════
# MAPEAMENTO DE NOMES DE PDV → NOME CANÔNICO
# Chaves em lowercase para match case-insensitive
# ══════════════════════════════════════════════════════════════════
PDV_NOME_MAP = {
    # Por fragmento de nome (match parcial)
    "barra funda":          "11# Extra Barra Funda",
    "glicério":             "5# Extra Glicério",
    "glicerio":             "5# Extra Glicério",
    "2a torre":             "5# Extra Glicério",
    "jacu":                 "8# Lopes - Jacu Pessego",
    "jacu pessego":         "8# Lopes - Jacu Pessego",
    "palmeiras":            "4# Extra Palmeiras",
    "rego freitas":         "2# Extra Rego Freitas",
    "são bernardo":         "14# Extra São Bernardo",
    "sao bernardo":         "14# Extra São Bernardo",
    "são caetano":          "13# Extra São Caetano",
    "sao caetano":          "13# Extra São Caetano",
    # Por MAC/ID exato (VendTEF/PayBlu)
    "tef 78:21:84:ee:80:c6": "34# Gestão You Go Vila Mariana",
    "tef 78:21:84:ee:8a:fe": "34# Gestão You Go Vila Mariana",
    "tef b8:d6:1a:83:2e:7a": "34# Gestão You Go Vila Mariana",
    "tef b8:d6:1a:83:2f:5e": "4# Gestão You Go Vila Mariana",
    "tef b8:d6:1a:83:2f:6a": "34# Gestão You Go Vila Mariana",
    "tef b8:d6:1a:83:2f:86": "34# Gestão You Go Vila Mariana",
    "78:21:84:ee:80:c6":     "34# Gestão You Go Vila Mariana",
    "78:21:84:ee:8a:fe":     "34# Gestão You Go Vila Mariana",
    "b8:d6:1a:83:2e:7a":     "34# Gestão You Go Vila Mariana",
    "b8:d6:1a:83:2f:5e":     "4# Gestão You Go Vila Mariana",
    "b8:d6:1a:83:2f:6a":     "34# Gestão You Go Vila Mariana",
    "b8:d6:1a:83:2f:86":     "34# Gestão You Go Vila Mariana",
}

def normalizar_pdv(nome_raw):
    """Normaliza nome de PDV usando PDV_NOME_MAP. Match exato primeiro, depois parcial."""
    if not nome_raw:
        return nome_raw
    s = str(nome_raw).strip()
    sl = s.lower()
    # 1. Match exato
    if sl in PDV_NOME_MAP:
        return PDV_NOME_MAP[sl]
    # 2. Match parcial (fragmento contido no nome)
    for fragmento, canonico in PDV_NOME_MAP.items():
        if fragmento in sl:
            return canonico
    return s



def normalizar_nsu(nsu_raw):
    """Converte NSU em notação científica (ex: '6,15212E+11') para inteiro string ('615212000000')."""
    if not nsu_raw:
        return nsu_raw
    s = str(nsu_raw).strip()
    if not s or s in ('-', '–', 'Não Informado', 'None'):
        return s
    import re as _re
    if _re.search(r'[Ee][+\-]\d+', s):
        try:
            val = float(s.replace(',', '.'))
            return str(int(val))
        except Exception:
            return s
    return s



async def dispensar_modal(page):
    """Detecta e fecha o modal 'Novos tutoriais' se ele estiver visível."""
    seletores_fechar = [
        "button#btn_close",
        ".closeModalNovidade",
        "button:has-text('Fechar')",
        ".modal-footer button",
    ]
    for sel in seletores_fechar:
        try:
            if await page.locator(sel).is_visible():
                log.info(f"Dispensando modal usando o seletor: {sel}")
                await page.click(sel)
                await asyncio.sleep(1)
                return True
        except Exception:
            continue
    return False


async def fazer_login(page) -> bool:
    """Preenche as credenciais de login no ERP Vending. Tenta até 3 vezes."""
    for tentativa in range(1, 4):
        try:
            log.info(f"Login ERP: tentativa {tentativa}/3 — URL atual: {page.url}")

            # Verifica se já está logado (navbar presente)
            try:
                if await page.locator("a#navbarVendtef, #navbarPayblu, .navbar-brand").first.is_visible(timeout=3000):
                    log.info("ERP: sessão já autenticada, pulando login.")
                    return True
            except Exception:
                pass

            seletores_usuario = [
                "#username",
                "input[name='username']",
                "input[name='user']",
                "input[type='text']",
            ]
            campo_usuario = None
            for sel in seletores_usuario:
                try:
                    await page.wait_for_selector(sel, timeout=5000)
                    campo_usuario = sel
                    break
                except PWTimeout:
                    continue

            if not campo_usuario:
                log.warning(f"Login ERP tentativa {tentativa}: campo de usuário não encontrado. URL: {page.url}")
                if tentativa < 3:
                    await page.goto("https://www.erpvending.com.br/", timeout=TIMEOUT_MS, wait_until="domcontentloaded")
                    await asyncio.sleep(3)
                    continue
                return False

            await page.fill(campo_usuario, LOGIN)

            seletores_senha = [
                "#password",
                "input[type='password']",
            ]
            campo_senha = None
            for sel in seletores_senha:
                try:
                    await page.wait_for_selector(sel, timeout=3000)
                    await page.fill(sel, SENHA)
                    campo_senha = sel
                    break
                except PWTimeout:
                    continue

            if not campo_senha:
                log.warning(f"Login ERP tentativa {tentativa}: campo de senha não encontrado.")
                if tentativa < 3:
                    await asyncio.sleep(3)
                    continue
                return False

            seletores_submit = [
                "input#login",
                "#login",
                "button[type='submit']",
                "input[type='submit']",
            ]
            botao_clicked = False
            for sel in seletores_submit:
                try:
                    await page.wait_for_selector(sel, timeout=3000)
                    await page.click(sel)
                    botao_clicked = True
                    break
                except Exception:
                    continue

            if not botao_clicked:
                log.warning(f"Login ERP tentativa {tentativa}: botão de submit não encontrado.")
                if tentativa < 3:
                    await asyncio.sleep(3)
                    continue
                return False

            await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
            await asyncio.sleep(2)

            # Verifica se o login foi bem-sucedido (navbar aparece)
            try:
                await page.wait_for_selector("a#navbarVendtef, #navbarPayblu, .navbar-nav", timeout=8000)
                log.info("Login no ERP Vending realizado com sucesso.")
                await dispensar_modal(page)
                return True
            except Exception:
                log.warning(f"Login ERP tentativa {tentativa}: navbar não apareceu após submit. URL: {page.url}")
                # Verifica se há mensagem de erro na página
                try:
                    err_text = await page.locator(".alert, .error, #error, .login-error").first.inner_text(timeout=2000)
                    log.warning(f"Mensagem de erro na página de login: {err_text[:200]}")
                except Exception:
                    pass
                if tentativa < 3:
                    await page.goto("https://www.erpvending.com.br/", timeout=TIMEOUT_MS, wait_until="domcontentloaded")
                    await asyncio.sleep(4)
                    continue
                return False

        except Exception as e:
            log.error(f"Login ERP tentativa {tentativa} — erro inesperado: {e}")
            if tentativa < 3:
                await asyncio.sleep(4)
                continue
            return False

    return False


class AppsScriptRedirectHandler(urllib.request.HTTPRedirectHandler):
    def http_error_302(self, req, fp, code, msg, headers):
        new_url = headers.get('Location')
        new_req = urllib.request.Request(new_url, headers={"User-Agent": "Mozilla/5.0"})
        return self.parent.open(new_req)
        
    http_error_301 = http_error_303 = http_error_307 = http_error_302


def sincronizar_csv_local():
    """Baixa o CSV completo do Google Sheets e salva nos arquivos locais (CSV e JS)."""
    csv_file_files = Path(__file__).parent / "relatorio_vendas_geral.csv"
    csv_file_cunha = Path(__file__).parent.parent / "cunha gago" / "relatorio_vendas_geral.csv"
    js_file_files = Path(__file__).parent / "relatorio_vendas_geral.js"
    js_file_cunha = Path(__file__).parent.parent / "cunha gago" / "relatorio_vendas_geral.js"
    
    try:
        log.info("Sincronizando CSV/JS local a partir do Google Sheets...")
        csv_bytes = None
        # Tenta a URL pública direta (muito mais rápido e não dá timeout)
        url_pub = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSQ6-Y8FO-Zspn95dDGPARIHVexMJ5foVWU4vsIuq_OdpAr0k6JY2OwrS2VNfygSkJ6dT1aXaPx6uuq/pub?gid=40793617&single=true&output=csv"
        try:
            req = urllib.request.Request(url_pub, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as r:
                csv_bytes = r.read()
            log.info("CSV sincronizado via link público do Google Sheets com sucesso.")
        except Exception as pub_err:
            log.warning(f"Falha ao baixar do link público ({pub_err}). Tentando Apps Script...")
            req = urllib.request.Request(
                APPS_SCRIPT_URL + "?gid=40793617",
                headers={"User-Agent": "Mozilla/5.0"}
            )
            with urllib.request.urlopen(req, timeout=120) as r:
                csv_bytes = r.read()
        
        if not csv_bytes:
            raise ValueError("Não foi possível carregar dados do Google Sheets.")
            
        lines = csv_bytes.decode("utf-8").splitlines()
        log.info(f"CSV sincronizado: {len(lines)} linhas do Google Sheets.")
        
        # Gerar o JS content
        csv_text = csv_bytes.decode("utf-8")
        escaped_csv_text = csv_text.replace("`", "\\`").replace("${", "\\${")
        js_content = f"window.LAVAI_CSV_DATA = `\n{escaped_csv_text}`;\n"
        js_bytes = js_content.encode("utf-8")
        
        for csv_file, js_file in [(csv_file_files, js_file_files), (csv_file_cunha, js_file_cunha)]:
            try:
                csv_file.parent.mkdir(parents=True, exist_ok=True)
                with open(csv_file, "wb") as f:
                    f.write(b'\xef\xbb\xbf')  # BOM UTF-8
                    f.write(csv_bytes)
                log.info(f"CSV local atualizado: {csv_file}")
                
                with open(js_file, "wb") as f:
                    f.write(b'\xef\xbb\xbf')  # BOM UTF-8
                    f.write(js_bytes)
                log.info(f"JS local atualizado: {js_file}")
            except Exception as e:
                log.error(f"Erro ao salvar arquivos locais em {csv_file.parent}: {e}")
    except Exception as e:
        log.warning(f"Nao foi possivel sincronizar CSV/JS local do Sheets: {e}")


def enviar_e_salvar_vendas(rows, sync_csv=True):
    """
    Envia via POST para o Google Apps Script (deduplicação é feita no servidor).
    Se sync_csv=True, sincroniza o CSV completo do Sheets para os arquivos locais após envio.
    """
    if not rows:
        log.warning("Nenhuma transacao para salvar ou enviar.")
        return

    log.info(f"Processando {len(rows)} transacoes...")

    # Enviar via POST para Google Apps Script
    if not APPS_SCRIPT_URL or "SUA_URL" in APPS_SCRIPT_URL:
        log.warning("APPS_SCRIPT_URL nao configurado no robo. Envio para Google Sheets ignorado.")
        return

    try:
        # Enviar as linhas no formato de 22 colunas
        payload = {
            "gid": "40793617",
            "rows": rows
        }
        data = json.dumps(payload).encode("utf-8")
        
        opener = urllib.request.build_opener(AppsScriptRedirectHandler())
        req = urllib.request.Request(
            APPS_SCRIPT_URL,
            data=data,
            headers={
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0"
            },
            method="POST"
        )
        
        log.info("Enviando dados para o Google Sheets via Apps Script...")
        with opener.open(req, timeout=90) as response:
            res_data = response.read().decode("utf-8")
            if res_data.strip().startswith("<!DOCTYPE html>"):
                if "doPost" in res_data:
                    log.error("Erro no Apps Script: A funcao 'doPost' nao foi encontrada.")
                else:
                    log.error("Erro no Apps Script: O script retornou erro HTML.")
                return

            res_json = json.loads(res_data)
            if res_json.get("status") == "success":
                log.info(f"Planilha atualizada! {res_json.get('inserted')} novas transacoes adicionadas.")
                # Sincronizar o CSV local com o estado completo do Google Sheets
                if sync_csv:
                    sincronizar_csv_local()
            elif res_json.get("status") == "error":
                log.error(f"Erro retornado do Apps Script: {res_json.get('message')}")
            else:
                log.error(f"Resposta inesperada do Apps Script: {res_json}")
    except Exception as e:
        log.error(f"Erro ao enviar dados para a planilha: {e}")


def get_field_case_insensitive(tx, keys, default=""):
    if not isinstance(tx, dict):
        return default
    for k, v in tx.items():
        if k.lower() in keys:
            if v is not None:
                return v
    return default


def map_vmpay_to_csv_row(tx):
    cliente = "Estoque - LAVAÍ"
    
    # 1. Maquina (PDV Location) — API retorna location: {"name": "PDV X"} e machine: {"asset_number": "X"}
    location_obj = get_field_case_insensitive(tx, ["location"])
    if isinstance(location_obj, dict):
        maquina = get_field_case_insensitive(location_obj, ["name", "description"]) or ""
    else:
        maquina = ""
    if not maquina:
        mach_obj = get_field_case_insensitive(tx, ["machine", "vending_machine"])
        if isinstance(mach_obj, dict):
            maquina = get_field_case_insensitive(mach_obj, ["asset_number", "name", "description"]) or ""
    if not maquina:
        maquina = get_field_case_insensitive(tx, ["client", "client_name", "cliente", "local", "location_name", "ponto_de_venda", "pdv"]) or "VMPay - Máquina"
        
    modelo = get_field_case_insensitive(tx, ["machine_model", "modelo", "modelo de máquina", "modelo de maquina"]) or ""
    fabricante = "VMPay"
    
    # 2. Pagamento
    pagamento = get_field_case_insensitive(tx, ["payment_method", "payment", "meio_pagamento", "pagamento"])
    if isinstance(pagamento, dict):
        pagamento = get_field_case_insensitive(pagamento, ["method", "name", "description"])
    if not pagamento:
        pagamento = "VMPay"
        
    # 3. Produto — API retorna good: {"name": "Lavadora"} ou {"name": "Secadora"}
    good_obj = get_field_case_insensitive(tx, ["good"])
    if isinstance(good_obj, dict):
        produto = get_field_case_insensitive(good_obj, ["name", "description"]) or "Indefinido"
    else:
        produto = get_field_case_insensitive(tx, ["product_name", "product", "item", "produto"])
        if isinstance(produto, dict):
            produto = get_field_case_insensitive(produto, ["name", "description"])
        if not produto:
            produto = "Indefinido"
        
    # 4. Mola ID (Coil)
    mola_id = get_field_case_insensitive(tx, ["coil_id", "coil", "canaleta", "mola", "mola_id"])
    if isinstance(mola_id, dict):
        mola_id = get_field_case_insensitive(mola_id, ["number", "id", "code"])
    mola_id = str(mola_id).strip()
    
    # 5. Value
    total_r = get_field_case_insensitive(tx, ["transaction_value", "amount", "value", "total", "valor", "valor_venda", "valor_total", "price", "preco", "valor (r$)"], 0.0)
    try:
        if isinstance(total_r, (int, float)):
            total_r = float(total_r)
        else:
            val_str = str(total_r).strip().replace("R$", "").strip()
            if "," in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            total_r = float(val_str)
            
        if total_r >= 100 and float(total_r) == int(total_r):
            total_r = total_r / 100.0
    except Exception:
        total_r = 0.0
        
    venda_r = total_r
    preco_r = "Não Informado"
    cod_promocional = "Não utilizado"
    
    # 6. Date & Time
    dt_str = get_field_case_insensitive(tx, ["occurred_at", "created_at", "date", "sale_date", "data", "data_hora", "data/hora"])
    data_br = ""
    hora_br = ""
    if dt_str:
        try:
            dt_str = str(dt_str).strip()
            if 'T' in dt_str:
                # API VMPay retorna UTC — converter para UTC-3 (horário de SP)
                from datetime import datetime as _dt, timezone as _tz, timedelta as _td
                _utc = _tz(timedelta(hours=0))
                _sp  = _tz(timedelta(hours=-3))
                # Remove 'Z' ou offset se houver
                dt_clean = dt_str.replace('Z', '+00:00')
                try:
                    dt_obj = _dt.fromisoformat(dt_clean).astimezone(_sp)
                except Exception:
                    # fallback: assume UTC puro
                    raw = dt_str.split('T')
                    d_parts = raw[0].split('-')
                    t_parts = raw[1][:8].split(':')
                    dt_obj = _dt(int(d_parts[0]), int(d_parts[1]), int(d_parts[2]),
                                 int(t_parts[0]), int(t_parts[1]), int(t_parts[2]),
                                 tzinfo=_tz(timedelta(hours=0))).astimezone(_tz(timedelta(hours=-3)))
                data_br = dt_obj.strftime("%d/%m/%Y")
                hora_br = dt_obj.strftime("%H:%M:%S")
            elif ' ' in dt_str:
                date_part, time_part = dt_str.split(' ', 1)
                data_br = date_part
                hora_br = time_part[:8]
            else:
                data_br = dt_str
        except Exception:
            data_br = str(dt_str)
            
    n_logico = get_field_case_insensitive(tx, ["logical_number", "serial", "n_logico", "nlogico", "vmbox"])
    if not n_logico:
        mach_obj = get_field_case_insensitive(tx, ["machine", "vending_machine"])
        if isinstance(mach_obj, dict):
            n_logico = get_field_case_insensitive(mach_obj, ["logical_number", "serial", "vmbox"])
    if not n_logico:
        n_logico = ""
        
    nsu = get_field_case_insensitive(tx, ["request_number", "nsu", "transaction_id", "id", "requisicao", "requisicao_id", "uuid"])
    auth = get_field_case_insensitive(tx, ["authorization_code", "auth", "autorizacao", "codigo_autorizacao", "código de autorização"])
    
    tipo_cartao = get_field_case_insensitive(tx, ["card_type", "type", "tipo_cartao", "tipo de cartão", "eft_card_type"])
    if isinstance(tipo_cartao, dict):
        tipo_cartao = get_field_case_insensitive(tipo_cartao, ["name", "description"])
        
    rede = get_field_case_insensitive(tx, ["network", "rede", "provedor", "adquirente", "eft_provider", "eft_authorizer"])
    if isinstance(rede, dict):
        rede = get_field_case_insensitive(rede, ["name", "description"])
        
    bandeira = get_field_case_insensitive(tx, ["card_brand", "brand", "bandeira", "cartao", "eft_card_brand"])
    if isinstance(bandeira, dict):
        bandeira = get_field_case_insensitive(bandeira, ["name", "description"])
        
    usuario = get_field_case_insensitive(tx, ["operator_name", "usuario", "consumidor", "usuario_nome"])
    no_cartao = get_field_case_insensitive(tx, ["card_number", "no_cartao", "numero_cartao", "número do cartão"])
    matricula = ""
    
    return [
        str(cliente).strip(), str(maquina).strip(), str(modelo).strip(), str(fabricante).strip(), 
        str(pagamento).strip(), str(produto).strip(), str(mola_id).strip(), 
        str(venda_r).replace('.', ','), str(preco_r), str(total_r).replace('.', ','),
        str(cod_promocional), str(data_br).strip(), str(hora_br).strip(), str(n_logico).strip(), 
        str(nsu).strip(), str(auth).strip(), str(tipo_cartao).strip(), str(rede).strip(), 
        str(bandeira).strip(), str(usuario).strip(), str(no_cartao).strip(), str(matricula).strip()
    ]


def clean_header_string(s):
    import unicodedata
    if s is None:
        return ""
    s = str(s).lower().strip()
    # Remove accents
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Remove non-alphanumeric
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


def fix_encoding(s):
    if not isinstance(s, str):
        return s
    try:
        return s.encode('latin-1', errors='replace').decode('utf-8', errors='ignore')
    except Exception:
        return s


def map_sq_location(location):
    fixed_loc = fix_encoding(location)
    cleaned_loc = clean_header_string(fixed_loc)
    
    SQ_LOCATION_MAP = {
        "Lavaí – BARRA FUNDA": "11# Extra Barra Funda",
        "Lavaí – Glicerio": "5# Extra Glicério",
        "Lavaí - Itaquera EXTRA": "1# Extra Itaquera",
        "Lavaí - Maria Luiza DIA": "7# Dia - Maria Luiza",
        "Lavaí - Rego Freitas EXTRA": "2# Extra Rego Freitas",
        "Lavaí - Rio Branco": "3# Extra Rio Branco",
        "Lavaí - Rua das Palmeiras EXTRA": "4# Extra Palmeiras",
        "Lavaí - Rua das Palmeiras EXTRA ": "4# Extra Palmeiras",
        "Lavaí - Vila Mariana": "10# Extra Vila Mariana"
    }
    
    for k, v in SQ_LOCATION_MAP.items():
        if clean_header_string(k) == cleaned_loc:
            return v
    return None


def find_header_index(headers, search_options):
    # Try exact match first
    for opt in search_options:
        opt_clean = clean_header_string(opt)
        for idx, h in enumerate(headers):
            if h is None:
                continue
            h_clean = clean_header_string(h)
            if opt_clean == h_clean:
                return idx
    # Fallback to opt in header
    for opt in search_options:
        opt_clean = clean_header_string(opt)
        for idx, h in enumerate(headers):
            if h is None:
                continue
            h_clean = clean_header_string(h)
            if opt_clean in h_clean:
                return idx
    return None


def resolve_excel_headers(headers):
    mappings = {
        "data_hora": ["data/hora", "data", "date"],
        "pdv": ["pdv", "ponto de venda", "localidade", "cliente"],
        "local": ["local"],
        "local_interno": ["local interno", "ponto de captura", "coil"],
        "modelo_maquina": ["modelo de máquina", "modelo de maquina", "modelo"],
        "vmbox": ["vmbox", "número lógico", "numero logico", "n lógico", "n logico"],
        "uuid": ["uuid"],
        "tipo": ["tipo"],
        "produto": ["produto"],
        "provedor": ["provedor"],
        "adquirente": ["adquirente"],
        "cartao": ["cartão", "cartao"],
        "tipo_cartao": ["tipo de cartão", "tipo de cartao"],
        "numero_cartao": ["número do cartão", "numero do cartao"],
        "valor": ["valor (R$)", "valor", "total"],
        "requisicao": ["requisi", "nsu"],
        "autorizacao": ["autoriza", "auth"],
        "consumidor": ["consumidor"],
        "cliente": ["cliente"],
        "autorizador": ["autorizador"]
    }
    resolved = {}
    for key, opts in mappings.items():
        resolved[key] = find_header_index(headers, opts)
    return resolved


def find_header_row_and_values(sheet):
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row and row[0]:
            row_str = str(row[0]).strip().lower()
            if "data" in row_str or "date" in row_str:
                return r_idx + 1, row
    return 1, None


def is_recent_date(dt_val, now):
    current_year = now.year
    current_month = now.month
    
    if current_month == 1:
        prev_year = current_year - 1
        prev_month = 12
    else:
        prev_year = current_year
        prev_month = current_month - 1
        
    if isinstance(dt_val, datetime):
        return (dt_val.year == current_year and dt_val.month == current_month) or \
               (dt_val.year == prev_year and dt_val.month == prev_month)
               
    dt_str = str(dt_val).strip()
    m_curr_str = f"/{current_month:02d}/{current_year}"
    m_prev_str = f"/{prev_month:02d}/{prev_year}"
    
    y_curr_str = f"{current_year}-{current_month:02d}"
    y_prev_str = f"{prev_year}-{prev_month:02d}"
    
    return m_curr_str in dt_str or m_prev_str in dt_str or y_curr_str in dt_str or y_prev_str in dt_str


def map_excel_row_to_csv_row(row, resolved):
    def get_val(key):
        idx = resolved.get(key)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    cliente = "Estoque - LAVAÍ"
    
    # 1. Maquina (PDV Location)
    maquina = get_val("pdv") or get_val("local") or get_val("cliente") or "VMPay - Máquina"
    
    modelo = get_val("modelo_maquina") or ""
    fabricante = "VMPay"
    
    # 2. Pagamento
    tipo = str(get_val("tipo") or "").lower()
    autorizador = str(get_val("autorizador") or "").lower()
    
    if "pix" in tipo or "pix" in autorizador:
        pagamento = "PIX"
    elif "tef" in tipo or "cart" in tipo or "pinpad" in tipo:
        pagamento = "TEF"
    else:
        pagamento = "TEF"
        
    # 3. Produto
    produto = get_val("produto") or "Indefinido"
    
    # 4. Mola ID (Coil)
    mola_id = get_val("local_interno") or get_val("local") or ""
    
    # 5. Value
    val_raw = get_val("valor") or 0.0
    total_r = 0.0
    try:
        if isinstance(val_raw, (int, float)):
            total_r = float(val_raw)
        else:
            val_str = str(val_raw).strip().replace("R$", "").strip()
            if "," in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            total_r = float(val_str)
    except Exception:
        total_r = 0.0
        
    venda_r = total_r
    preco_r = "Não Informado"
    cod_promocional = "Não utilizado"
    
    # 6. Date & Time
    dt_val = get_val("data_hora")
    data_br = ""
    hora_br = ""
    if dt_val:
        try:
            if isinstance(dt_val, datetime):
                data_br = dt_val.strftime("%d/%m/%Y")
                hora_br = dt_val.strftime("%H:%M:%S")
            else:
                dt_str = str(dt_val).strip()
                if ' ' in dt_str:
                    date_part, time_part = dt_str.split(' ', 1)
                    if '-' in date_part:
                        parts = date_part.split('-')
                        if len(parts) == 3:
                            data_br = f"{parts[2]}/{parts[1]}/{parts[0]}"
                        else:
                            data_br = date_part
                    else:
                        data_br = date_part
                    hora_br = time_part[:8]
                else:
                    if '-' in dt_str:
                        parts = dt_str.split('-')
                        if len(parts) == 3:
                            data_br = f"{parts[2]}/{parts[1]}/{parts[0]}"
                        else:
                            data_br = dt_str
                    else:
                        data_br = dt_str
        except Exception:
            data_br = str(dt_val)
            
    n_logico = get_val("vmbox") or ""
    nsu = get_val("requisicao") or get_val("uuid") or ""
    auth = get_val("autorizacao") or ""
    tipo_cartao = get_val("tipo_cartao") or ""
    rede = get_val("provedor") or get_val("adquirente") or ""
    bandeira = get_val("cartao") or ""
    usuario = get_val("consumidor") or get_val("cliente") or ""
    no_cartao = get_val("numero_cartao") or ""
    matricula = ""
    
    return [
        str(cliente).strip(), str(maquina).strip(), str(modelo).strip(), str(fabricante).strip(), 
        str(pagamento).strip(), str(produto).strip(), str(mola_id).strip(), 
        str(venda_r).replace('.', ','), str(preco_r), str(total_r).replace('.', ','),
        str(cod_promocional), str(data_br).strip(), str(hora_br).strip(), str(n_logico).strip(), 
        str(nsu).strip(), str(auth).strip(), str(tipo_cartao).strip(), str(rede).strip(), 
        str(bandeira).strip(), str(usuario).strip(), str(no_cartao).strip(), str(matricula).strip()
    ]


def coletar_vmpay_api():
    log.info("Buscando dados de transações cashless da API VMPay...")
    
    # Verifica se já temos dados locais salvos em algum lugar
    tem_dados = False
    for path in [CSV_VMPAY_LOCAL, Path(__file__).parent / "vmpay_local.js", Path(__file__).parent.parent / "vmpay_local.js"]:
        if path.exists() and len(path.read_text(encoding="utf-8")) > 1000:
            tem_dados = True
            break

    now = datetime.now()
    if tem_dados:
        # Se temos dados, busca apenas os últimos 30 dias para uma atualização rápida
        start_date = now - timedelta(days=30)
        log.info("Dados locais de VMPay encontrados. Buscando apenas os últimos 30 dias de transações para atualização rápida...")
    else:
        # Se não temos dados, busca desde o início do ano
        start_date = datetime(2026, 1, 1, 0, 0, 0)
        log.info("Dados locais de VMPay não encontrados ou vazios. Buscando desde o início do ano (2026-01-01)...")

    end_date = datetime(now.year, now.month, now.day, 23, 59, 59)
    start_date_str = start_date.strftime("%Y-%m-%dT%H:%M:%S")
    end_date_str = end_date.strftime("%Y-%m-%dT%H:%M:%S")
    
    log.info(f"API VMPay: Filtro de período de {start_date_str} até {end_date_str}")
    
    import urllib.parse
    
    # Otimizado para trazer o limite máximo permitido (1000 por página)
    per_page = 1000
    page = 1
    has_more = True
    all_txs = []
    
    while has_more:
        api_url = f"{VMPAY_API_URL}?access_token={VMPAY_TOKEN}&start_date={urllib.parse.quote(start_date_str)}&end_date={urllib.parse.quote(end_date_str)}&per_page={per_page}&page={page}&contentType=json"
        log.info(f"Buscando página {page} da API VMPay (url: {api_url})...")
        req = urllib.request.Request(api_url, headers={"User-Agent": "Mozilla/5.0"})
        
        try:
            with urllib.request.urlopen(req, timeout=40) as r:
                res_data = r.read().decode("utf-8")
                data = json.loads(res_data)
                
                tx_list = []
                if isinstance(data, list):
                    tx_list = data
                elif isinstance(data, dict):
                    for key in ("transactions", "data", "results", "items"):
                        if key in data and isinstance(data[key], list):
                            tx_list = data[key]
                            break
                
                if tx_list:
                    all_txs.extend(tx_list)
                    log.info(f"API VMPay: página {page} — {len(tx_list)} transações.")
                    if len(tx_list) < per_page:
                        has_more = False
                    else:
                        page += 1
                else:
                    log.info(f"API VMPay: página {page} vazia ou fim dos dados.")
                    has_more = False
                    break
        except urllib.error.HTTPError as e:
            log.error(f"Erro HTTP ao chamar a API VMPay na página {page} ({e.code} {e.reason}). Abortando coleta VMPay.")
            if e.code == 401:
                log.warning("O token de acesso da API VMPay está expirado ou é inválido (401 Unauthorized).")
            return None
        except Exception as e:
            log.error(f"Erro geral ao chamar a API VMPay na página {page} ({e}). Abortando coleta VMPay.")
            return None
            
    log.info(f"API VMPay: total de {len(all_txs)} transações coletadas no período.")
    rows = []
    for tx in all_txs:
        if isinstance(tx, dict):
            rows.append(map_vmpay_to_csv_row(tx))
    return rows


def is_june_2026(dt_val):
    if not dt_val:
        return False
    if isinstance(dt_val, datetime):
        return dt_val.year == 2026 and dt_val.month == 6
    dt_str = str(dt_val).strip()
    return "/06/2026" in dt_str or "2026-06" in dt_str

def obter_caminho_excel(nome_arquivo):
    def normalizar(s):
        import re
        s = s.lower()
        if s.startswith("temp_read_only_"):
            s = s[len("temp_read_only_"):]
        s = re.sub(r'[^a-z0-9]', '', s)
        return s

    nome_norm = normalizar(nome_arquivo)

    # Pastas de busca
    pastas = [
        Path(r"C:\Users\badad\OneDrive\Desktop\gateway LAVAI"),
        Path(__file__).parent / "kpi",
        Path(__file__).parent
    ]

    caminhos_existentes = []
    for pasta in pastas:
        if not pasta.exists():
            continue
        try:
            for p in pasta.glob("*.xlsx"):
                if normalizar(p.name) == nome_norm:
                    caminhos_existentes.append(p)
        except Exception:
            pass

    if not caminhos_existentes:
        return None

    # Retorna o arquivo mais recentemente modificado entre todos os encontrados
    novo_caminho = max(caminhos_existentes, key=lambda p: p.stat().st_mtime)
    log.info(f"obter_caminho_excel('{nome_arquivo}'): selecionado o mais recente: {novo_caminho} (Modificado em: {datetime.fromtimestamp(novo_caminho.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')})")
    return novo_caminho

def map_vmpay_excel_row(row, headers_idx):
    def get(col_name):
        idx = headers_idx.get(col_name)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    data_hora = get('Data/hora')
    if " " in data_hora:
        data_br, hora_br = data_hora.split(" ", 1)
    else:
        data_br = data_hora
        hora_br = "12:00:00"

    cliente = "Estoque - LAVAÍ"
    maquina = get('PDV') or get('Local') or "VMPay - Máquina"
    modelo = get('Modelo de máquina')
    fabricante = "VMPay"
    pagamento = "CASHLESS"
    produto = get('Produto')
    mola_id = ""
    total_r = get('Valor (R$)')
    try:
        val_float = float(total_r)
        total_r = f"{val_float:.2f}".replace('.', ',')
        venda_r = total_r
    except ValueError:
        venda_r = "0,00"
        total_r = "0,00"

    preco_r = "Não Informado"
    cod_promocional = "Não utilizado"
    n_logico = get('VMbox')
    nsu = get('Requisição') or get('Uuid')
    auth = get('Código de Autorização')
    tipo_cartao = get('Tipo de cartão')
    rede = get('Provedor') or get('Adquirente') or "VMPay"
    bandeira = get('Cartão')
    usuario = get('Consumidor')
    no_cartao = get('Número do cartão')
    matricula = ""

    return [
        cliente, maquina, modelo, fabricante, pagamento, produto, mola_id,
        venda_r, preco_r, total_r, cod_promocional, data_br, hora_br, n_logico,
        nsu, auth, tipo_cartao, rede, bandeira, usuario, no_cartao, matricula
    ]

def coletar_vmpay_excel():
    master_path = obter_caminho_excel("Vmpay 2025.xlsx")
    all_rows = []
    if not master_path or not master_path.exists():
        log.warning("Arquivo Vmpay 2025.xlsx nao encontrado.")
        return all_rows

    log.info(f"Lendo dados de VMPay 2025 do arquivo: {master_path}")
    try:
        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
        sheet = wb.active

        headers = []
        for row in sheet.iter_rows(max_row=1, values_only=True):
            headers = [str(h).strip() if h else '' for h in row]

        headers_idx = {h: i for i, h in enumerate(headers)}
        date_idx = headers_idx.get('Data/hora', 0)
        estado_idx = headers_idx.get('Estado')

        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 10:
                continue

            if estado_idx is not None and estado_idx < len(row):
                if str(row[estado_idx]).strip().upper() != 'OK':
                    continue

            dt_val = row[date_idx] if date_idx < len(row) else None
            if not dt_val or is_june_2026(dt_val):
                continue

            mapped_row = map_vmpay_excel_row(row, headers_idx)
            all_rows.append(mapped_row)
            count += 1

        log.info(f"Processadas {count} transacoes de VMPay 2025 do arquivo.")
    except Exception as e:
        log.error(f"Erro ao ler arquivo Vmpay 2025 {master_path}: {e}")

    return all_rows


def merge_and_deduplicate(portal_rows, api_rows):
    seen = set()
    merged = []
    
    for r in portal_rows + api_rows:
        if len(r) < 22:
            continue
        nsu = normalizar_nsu(r[14].strip())
        r[14] = nsu  # normaliza in-place antes de enviar
        r[1]  = normalizar_pdv(r[1])  # normaliza nome do PDV in-place
        auth = r[15].strip()
        maq = r[1].strip()
        dt = r[11].strip()
        hr = r[12].strip()
        val = r[9].strip()
        
        if nsu and nsu not in ("Não Informado", "–", ""):
            key = f"nsu:{nsu}"
        elif auth and auth not in ("–", ""):
            key = f"auth:{auth}"
        else:
            key = f"dt:{maq}|{dt}|{hr}|{val}"
            
        if key not in seen:
            seen.add(key)
            merged.append(r)
            
    return merged



def clean_val(val_str):
    if val_str is None:
        return 0.0
    s = str(val_str).replace("R$", "").replace(" ", "").strip()
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.find(".") < s.find(","):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def gerar_pseudo_linha(location, machine_type, machine, pagamento, value, data_br, hora_br, nsu, tipo_cartao, bandeira):
    cliente = "Estoque - LAVAÍ"
    modelo = "Speed Queen (Nova)"
    fabricante = "Speed Queen"
    mola_id = machine
    venda_r = value
    preco_r = "Não Informado"
    total_r = value
    cod_promocional = "Não utilizado"
    n_logico = f"SQ-{machine}"
    auth = nsu
    rede = "Speed Queen"
    usuario = ""
    no_cartao = ""
    matricula = ""
    
    return [
        str(cliente).strip(), str(location).strip(), str(modelo).strip(), str(fabricante).strip(), 
        str(pagamento).strip(), str("Lavagem (SQ)" if "Washer" in machine_type or "lavar" in machine_type.lower() else "Secagem (SQ)").strip(), str(mola_id).strip(), 
        str(venda_r).replace('.', ','), str(preco_r), str(total_r).replace('.', ','),
        str(cod_promocional), str(data_br).strip(), str(hora_br).strip(), str(n_logico).strip(), 
        str(nsu).strip(), str(auth).strip(), str(tipo_cartao).strip(), str(rede).strip(), 
        str(bandeira).strip(), str(usuario).strip(), str(no_cartao).strip(), str(matricula).strip()
    ]


def processar_csv_sq(csv_content, start_date):
    rows = []
    
    parts = start_date.split("-")
    year_month = f"{parts[0]}{parts[1]}"
    
    reader = csv.reader(csv_content.splitlines())
    header_found = False
    for row in reader:
        if not row:
            continue
        if len(row) >= 10 and row[1] == "Location" and row[3] == "Machine":
            header_found = True
            continue
        if not header_found:
            continue
            
        if len(row) >= 10:
            location = row[1].strip()
            machine_type = row[2].strip()
            machine = row[3].strip()
            last_received = row[5].strip()
            coin_str = row[6].strip()
            mobile_str = row[9].strip()
            
            if not location or not machine or machine.startswith("Total") or machine.startswith("Grand"):
                continue
                
            mapped_location = map_sq_location(location)
            if not mapped_location:
                log.warning(f"Localizacao do SQInsights nao mapeada: {location}")
                continue
                
            data_br = ""
            hora_br = ""
            if last_received:
                try:
                    if " " in last_received:
                        dt_part, hr_part = last_received.split(" ", 1)
                        data_br = dt_part.strip()
                        hora_br = hr_part.strip() + ":00"
                    else:
                        data_br = last_received.strip()
                        hora_br = "12:00:00"
                except Exception:
                    data_br = last_received
                    hora_br = "12:00:00"
            else:
                data_br = datetime.now(FUSO_SP).strftime("%d/%m/%Y")
                hora_br = "12:00:00"
                
            ymd = ""
            if data_br:
                parts = data_br.split('/')
                if len(parts) == 3:
                    ymd = f"{parts[2]}{parts[1]}{parts[0]}"
            if not ymd:
                ymd = year_month
                
            loc_slug = clean_header_string(fix_encoding(location)).replace("lavai", "").strip()
            
            coin_val = clean_val(coin_str)
            if coin_val > 0:
                nsu = f"SQ-COIN-{loc_slug}-{machine}-{ymd}"
                rows.append(gerar_pseudo_linha(
                    location=mapped_location,
                    machine_type=machine_type,
                    machine=machine,
                    pagamento="MOEDA",
                    value=coin_val,
                    data_br=data_br,
                    hora_br=hora_br,
                    nsu=nsu,
                    tipo_cartao="Moeda",
                    bandeira="Moeda"
                ))
                
            mobile_val = clean_val(mobile_str)
            if mobile_val > 0:
                nsu = f"SQ-APP-{loc_slug}-{machine}-{ymd}"
                rows.append(gerar_pseudo_linha(
                    location=mapped_location,
                    machine_type=machine_type,
                    machine=machine,
                    pagamento="APLICATIVO",
                    value=mobile_val,
                    data_br=data_br,
                    hora_br=hora_br,
                    nsu=nsu,
                    tipo_cartao="Aplicativo",
                    bandeira="Aplicativo"
                ))
                
    log.info(f"Geradas {len(rows)} transacoes de moeda/aplicativo do SQInsights.")
    return rows


async def coletar_sq_excel():
    master_path = obter_caminho_excel("SQI - Fichas e APP.xlsx")
    all_rows = []
    
    if master_path and master_path.exists():
        log.info(f"Lendo dados de SQInsights do arquivo mestre: {master_path}")
        try:
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            sheet = wb.active
            
            now = datetime.now(FUSO_SP)
            count = 0
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx == 0:
                    continue
                if not row or len(row) < 8:
                    continue
                    
                location = row[0]
                machine_type = fix_encoding(row[1])
                machine = str(row[2])
                last_received = row[4]
                coin_str = row[5]
                mobile_str = row[7]
                
                if not location or not machine or not last_received:
                    continue
                    
                mapped_location = map_sq_location(location)
                if not mapped_location:
                    continue
                    
                # Permitir dados de junho no SQInsights (conforme solicitado pelo usuário)
                # if is_june_2026(last_received):
                #     continue
                    
                if isinstance(last_received, datetime):
                    data_br = last_received.strftime("%d/%m/%Y")
                    hora_br = last_received.strftime("%H:%M:%S")
                    ymd = last_received.strftime("%Y%m%d")
                else:
                    dt_str = str(last_received).strip()
                    if " " in dt_str:
                        date_part, time_part = dt_str.split(" ", 1)
                        data_br = date_part.strip()
                        hora_br = time_part.strip()
                    else:
                        data_br = dt_str.strip()
                        hora_br = "12:00:00"
                    ymd = data_br.replace("/", "")
                    
                loc_slug = clean_header_string(fix_encoding(location)).replace("lavai", "").strip()
                
                coin_val = clean_val(coin_str)
                if coin_val > 0:
                    nsu = f"SQ-COIN-{loc_slug}-{machine}-{ymd}"
                    all_rows.append(gerar_pseudo_linha(
                        location=mapped_location,
                        machine_type=machine_type,
                        machine=machine,
                        pagamento="MOEDA",
                        value=coin_val,
                        data_br=data_br,
                        hora_br=hora_br,
                        nsu=nsu,
                        tipo_cartao="Moeda",
                        bandeira="Moeda"
                    ))
                    count += 1
                    
                mobile_val = clean_val(mobile_str)
                if mobile_val > 0:
                    nsu = f"SQ-APP-{loc_slug}-{machine}-{ymd}"
                    all_rows.append(gerar_pseudo_linha(
                        location=mapped_location,
                        machine_type=machine_type,
                        machine=machine,
                        pagamento="APLICATIVO",
                        value=mobile_val,
                        data_br=data_br,
                        hora_br=hora_br,
                        nsu=nsu,
                        tipo_cartao="Aplicativo",
                        bandeira="Aplicativo"
                    ))
                    count += 1
                    
            log.info(f"Processadas {count} transações recentes de SQInsights do arquivo mestre.")
            return all_rows
        except Exception as e:
            log.error(f"Erro ao ler arquivo mestre SQInsights {master_path}: {e}")
            
    log.warning("Arquivo mestre do SQInsights não encontrado. Usando fallback para API/CDP.")
    return await coletar_sq_api()


def coletar_vendpago_excel():
    """
    Le dados de VendPago (Credito Remoto/Cashless) do arquivo Excel especifico.
    """
    import datetime as dt_mod

    master_path = obter_caminho_excel("vendpago 2026.xlsx")
    all_rows = []

    if not master_path or not master_path.exists():
        log.warning("Arquivo VendPago nao encontrado.")
        return all_rows

    log.info(f"Lendo dados de VendPago do arquivo: {master_path}")

    def fmt_val(v, idx):
        if v is None:
            return ""
        if isinstance(v, datetime):
            if idx == 11:
                return v.strftime("%d/%m/%Y")
            elif idx == 12:
                return v.strftime("%H:%M:%S")
            return str(v)
        if isinstance(v, dt_mod.time):
            return v.strftime("%H:%M:%S")
        if isinstance(v, (int, float)):
            if idx in (7, 8, 9):
                return str(v).replace('.', ',')
            return str(v)
        return str(v).strip()

    try:
        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
        sheet = wb.active

        header_row_num, header_values = find_header_row_and_values(sheet)
        if not header_values:
            header_values = next(sheet.iter_rows(values_only=True))
            header_row_num = 1

        headers_lower = [str(h).lower().strip() if h else '' for h in header_values]

        # Detecta layout VendPago Remoto: tem 'produto' na coluna 3 (antes de pagamento na coluna 4)
        is_credito_remoto = (
            len(headers_lower) > 4 and
            'produto' in (headers_lower[3] if len(headers_lower) > 3 else '') and
            'pagamento' in (headers_lower[4] if len(headers_lower) > 4 else '')
        )

        # Encontra indice da coluna data
        date_idx = 11
        for i, h in enumerate(headers_lower):
            if 'data' in h and 'hora' not in h:
                date_idx = i
                break

        count = 0

        for row in sheet.iter_rows(min_row=header_row_num + 1, values_only=True):
            if not row or len(row) < 10:
                continue

            dt_val = row[date_idx] if date_idx < len(row) else None
            if not dt_val or is_june_2026(dt_val):
                continue

            def get(i):
                return fmt_val(row[i], i) if i < len(row) else ""

            if is_credito_remoto:
                # Layout Credito Remoto: Cliente[0], PDV[1], Torre[2], Produto[3], Pagamento[4]
                # Remapeia: Fabricante='VendPago', Pagamento='CASHLESS', Produto->col5
                mapped_row = [
                    get(0),               # [0] Cliente
                    get(1),               # [1] Maquina (PDV)
                    "VendPago (Nova)",    # [2] Modelo
                    "VendPago",           # [3] Fabricante -> identifica fonte
                    "CASHLESS",           # [4] Pagamento -> detectado como VMPay/VendPago
                    get(3),               # [5] Produtos -> Produto do arquivo (Lavadora/Secadora)
                    "",                   # [6] Mola ID
                    get(9),               # [7] Venda R$
                    "Nao Informado",      # [8] Preco R$
                    get(9),               # [9] Total R$
                    "Nao utilizado",      # [10] Cod Promocional
                    get(11),              # [11] Data
                    get(12),              # [12] Hora
                    get(13),              # [13] N Logico
                    get(14),              # [14] NSU
                    get(15),              # [15] Autorizacao
                    get(16),              # [16] Tipo Cartao
                    get(17) or "VendPago",# [17] Rede
                    get(18),              # [18] Bandeira
                    get(19),              # [19] Usuario
                    get(20),              # [20] No Cartao
                    get(21),              # [21] Matricula
                ]
            else:
                # Layout portal TEF: mapeamento direto
                mapped_row = [fmt_val(row[i], i) if i < len(row) else "" for i in range(22)]

            all_rows.append(mapped_row)
            count += 1

        log.info(f"Processadas {count} transacoes recentes de VendPago do arquivo '{master_path.name}'.")

    except Exception as e:
        log.error(f"Erro ao ler arquivo VendPago {master_path}: {e}")

    return all_rows




def map_payblu_csv_row(row):
    """
    Mapeia uma linha do CSV PayBlu Private Label para o formato padrão de 22 colunas.
    Colunas CSV: Cliente[0]; Nº Serial[1]; MAC[2]; Matricula[3]; Nome Terminal[4];
                 Pagamento[5]; Produto[6]; Mola[7]; Preço[8]; Valor Pago[9];
                 Data[10]; Hora[11]; Usuário[12]
    """
    def g(i):
        return row[i].strip() if i < len(row) else ""

    cliente     = g(0) or "LAVAÍ - You Go"
    maquina     = g(4) or g(1)   # Nome Terminal ou Serial
    modelo      = "PayBlu (Private Label)"
    fabricante  = "PayBlu"
    pagamento   = g(5) or "PRIVATE LABEL"
    produto     = g(6) or "1 Pulso(s)"
    mola_id     = g(7)
    preco_r     = g(8) or "Não Informado"
    valor_raw   = g(9)
    data_br     = g(10)
    hora_br     = g(11)
    usuario     = g(12)

    # Valor: vem como "7" (inteiros em centavos) ou "7,00"
    total_r = 0.0
    try:
        v = valor_raw.replace("R$", "").replace(" ", "").strip()
        if "," in v:
            v = v.replace(".", "").replace(",", ".")
        total_r = float(v)
        # Se vier como inteiro (ex: "7") e for valor de R$7,00 → já está certo
        # Centavos só se >= 1000 e inteiro puro
        if total_r >= 1000 and total_r == int(total_r):
            total_r = total_r / 100.0
    except Exception:
        total_r = 0.0

    # NSU sintético único: PB-<Serial>|<Data>|<Hora> (sem NSU real no CSV)
    serial = g(1)
    nsu = f"PB-{serial}|{data_br}|{hora_br}" if serial else ""

    n_logico    = g(3)   # Matrícula como número lógico
    auth        = ""
    tipo_cartao = "Private Label"
    rede        = "PayBlu"
    bandeira    = "Private Label"
    no_cartao   = ""
    matricula   = g(3)

    return [
        cliente, maquina, modelo, fabricante,
        pagamento, produto, mola_id,
        str(total_r).replace(".", ","), preco_r, str(total_r).replace(".", ","),
        "Não utilizado", data_br, hora_br, n_logico,
        nsu, auth, tipo_cartao, rede,
        bandeira, usuario, no_cartao, matricula
    ]


async def _payblu_estabelecer_sso(page):
    """Estabelece sessão SSO no PayBlu via navbar do ERP. Retorna True se OK."""
    try:
        await page.goto(URL_ERP, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
        await dispensar_modal(page)

        payblu_sso_href = None
        for sel in ["a#navbarPayblu", "a#navbarpayblu", "a:has-text('PayBlu')", "a[href*='payblu']"]:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=3000):
                    payblu_sso_href = await el.get_attribute("href")
                    if payblu_sso_href:
                        log.info(f"PayBlu SSO: link encontrado via '{sel}'")
                        break
            except Exception:
                continue

        if payblu_sso_href:
            await page.goto(payblu_sso_href, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
            await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
            await asyncio.sleep(3)
            return True
        else:
            log.warning("PayBlu SSO: link não encontrado no navbar.")
            return False
    except Exception as e:
        log.error(f"PayBlu SSO: erro ao estabelecer sessão: {e}")
        return False


async def _payblu_baixar_mes(page, ano: int, mes: int, pasta_saida: Path) -> list:
    """
    Baixa o relatório PayBlu de um mês específico.
    Preenche os campos de data inicio/fim do formulário, clica Continuar → Download.
    Salva o CSV em pasta_saida/payblu_YYYY_MM.csv e retorna as rows mapeadas.
    """
    import calendar
    primeiro_dia = f"01/{mes:02d}/{ano}"
    ultimo_dia_n = calendar.monthrange(ano, mes)[1]
    ultimo_dia   = f"{ultimo_dia_n:02d}/{mes:02d}/{ano}"
    label_mes    = f"{ano}-{mes:02d}"

    log.info(f"PayBlu: coletando {label_mes} ({primeiro_dia} → {ultimo_dia})...")
    rows = []
    temp_csv = Path(__file__).parent / f"temp_payblu_{ano}_{mes:02d}.csv"

    try:
        await page.goto(URL_PAYBLU, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
        await asyncio.sleep(2)
        await dispensar_modal(page)

        # Verifica se não foi redirecionado para login
        if "login" in page.url.lower() or "erpvending" in page.url.lower():
            log.warning(f"PayBlu {label_mes}: redirecionado para login. Sessão perdida.")
            return rows

        # Inspeciona campos de data disponíveis no formulário
        campos_data = await page.locator("input[type='date'], input[name*='data'], input[name*='date'], input[name*='inicio'], input[name*='fim'], input[name*='start'], input[name*='end']").all()
        log.info(f"PayBlu {label_mes}: {len(campos_data)} campos de data encontrados.")

        if len(campos_data) >= 2:
            # Formulário com campos de data início e fim
            campo_inicio = campos_data[0]
            campo_fim    = campos_data[1]

            # Tenta preencher como input[type=date] (formato YYYY-MM-DD)
            tipo_inicio = await campo_inicio.get_attribute("type") or ""
            if tipo_inicio == "date":
                await campo_inicio.fill(f"{ano}-{mes:02d}-01")
                await campo_fim.fill(f"{ano}-{mes:02d}-{ultimo_dia_n:02d}")
            else:
                # Formato brasileiro dd/mm/yyyy
                await campo_inicio.click(click_count=3)
                await campo_inicio.type(primeiro_dia)
                await campo_fim.click(click_count=3)
                await campo_fim.type(ultimo_dia)

            log.info(f"PayBlu {label_mes}: datas preenchidas ({primeiro_dia} → {ultimo_dia}).")
        else:
            # Formulário sem filtro de data (retorna mês corrente) — só funciona para o mês atual
            now = datetime.now(FUSO_SP)
            if not (ano == now.year and mes == now.month):
                log.warning(f"PayBlu {label_mes}: formulário sem campos de data, pulando mês não-corrente.")
                return rows
            log.info(f"PayBlu {label_mes}: sem campos de data, baixando mês corrente.")

        # Clica Continuar
        await page.locator("input[value='Continuar'], input[type='submit'], button[type='submit']").first.click(timeout=10000)
        await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
        await asyncio.sleep(4)

        # Clica Download
        btn_locator = page.locator(
            "a:has-text('Download'), button:has-text('Download'), input[value='Download']"
        ).first

        async with page.expect_download(timeout=25000) as dl_info:
            await btn_locator.click()

        download = await dl_info.value
        await download.save_as(str(temp_csv))
        log.info(f"PayBlu {label_mes}: CSV baixado.")

        # Processa CSV
        with open(temp_csv, encoding="latin-1") as f:
            raw_lines = f.readlines()

        header_idx = 0
        for i, line in enumerate(raw_lines):
            if "Cliente" in line and "Pagamento" in line:
                header_idx = i
                break

        reader = csv.reader(raw_lines[header_idx + 1:], delimiter=";")
        count = 0
        for r in reader:
            if not r or not r[0].strip():
                continue
            if r[0].strip().startswith("Total"):
                continue
            r = [c.strip().strip('"') for c in r]
            if len(r) >= 11:
                rows.append(map_payblu_csv_row(r))
                count += 1

        log.info(f"PayBlu {label_mes}: {count} transações processadas.")

        # Salva arquivo mensal persistente
        if rows and pasta_saida:
            pasta_saida.mkdir(parents=True, exist_ok=True)
            arq_mes = pasta_saida / f"payblu_{ano}_{mes:02d}.csv"
            with open(arq_mes, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(["Cliente","Máquina","Modelo","Fabricante","Pagamento",
                                 "Produtos","Mola","Venda (R$)","Preço (R$)","Total",
                                 "Código Promocional","Data","Hora","Nº Logico","NSU",
                                 "Autorização","Tipo Cartão","Rede","Bandeira",
                                 "Usuário","Nº Cartão","Matricula"])
                for row in rows:
                    padded = list(row) + [""] * (22 - len(row))
                    writer.writerow(padded[:22])
            log.info(f"PayBlu {label_mes}: arquivo salvo em {arq_mes}.")

    except Exception as e:
        log.error(f"PayBlu {label_mes}: erro — {e}")
    finally:
        if temp_csv.exists():
            try: temp_csv.unlink()
            except: pass

    return rows


async def coletar_payblu(page):
    """
    Coleta relatórios PayBlu mês a mês para 2026 (jan até mês atual),
    salva arquivos mensais persistentes e retorna todas as rows consolidadas.
    """
    log.info("PayBlu: iniciando coleta mensal 2026...")
    todas_rows = []

    # Pasta para arquivos mensais
    pasta_payblu = Path(__file__).parent / "payblu_historico"

    # Estabelece SSO
    sso_ok = await _payblu_estabelecer_sso(page)
    if not sso_ok:
        log.warning("PayBlu: SSO falhou, tentando acesso direto...")

    now = datetime.now(FUSO_SP)
    ano_atual = now.year
    mes_atual = now.month

    # Coleta jan/2026 até mês corrente
    ano_inicio = 2026
    mes_inicio = 1

    for ano in range(ano_inicio, ano_atual + 1):
        m_inicio = mes_inicio if ano == ano_inicio else 1
        m_fim    = mes_atual  if ano == ano_atual  else 12
        for mes in range(m_inicio, m_fim + 1):
            # Verifica se já existe arquivo salvo para este mês (exceto mês atual — sempre atualiza)
            arq_mes = pasta_payblu / f"payblu_{ano}_{mes:02d}.csv"
            if arq_mes.exists() and not (ano == ano_atual and mes == mes_atual):
                log.info(f"PayBlu {ano}-{mes:02d}: arquivo já existe, carregando do disco.")
                rows_mes = _payblu_carregar_csv_mes(arq_mes)
                todas_rows.extend(rows_mes)
                continue

            # Precisa renovar SSO a cada ~3 meses para não expirar sessão
            if mes % 3 == 1 and mes > 1:
                await _payblu_estabelecer_sso(page)

            rows_mes = await _payblu_baixar_mes(page, ano, mes, pasta_payblu)
            todas_rows.extend(rows_mes)

            # Pequena pausa entre downloads para não sobrecarregar o servidor
            await asyncio.sleep(2)

    log.info(f"PayBlu: coleta concluída — {len(todas_rows)} transações no total.")
    return todas_rows


def _payblu_carregar_csv_mes(arq_path: Path) -> list:
    """Carrega rows de um arquivo CSV mensal salvo anteriormente."""
    rows = []
    try:
        with open(arq_path, encoding="utf-8", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            next(reader, None)  # pula cabeçalho
            for r in reader:
                if len(r) >= 22:
                    rows.append(r)
    except Exception as e:
        log.error(f"PayBlu: erro ao carregar {arq_path.name}: {e}")
    return rows


def coletar_yougo_excel():
    """
    Lê dados históricos do relatório Excel PayBlu You Go 2025.
    Colunas: Cliente[0] Serial[1] MAC[2] Matricula[3] NomeTerminal[4]
             Pagamento[5] Produto[6] Mola[7] Preco[8] ValorPago[9]
             Data[10] Hora[11] Usuario[12] NSU[13]
    """
    import datetime as _dt

    master_path = obter_caminho_excel("temp_read_only_Vendas_You_Go_25.xlsx")
    if not master_path:
        for nome in ["Vendas You Go 25.xlsx", "Vendas_You_Go_25.xlsx",
                     "vendas you go 25.xlsx", "vendas_you_go_25.xlsx"]:
            master_path = obter_caminho_excel(nome)
            if master_path:
                break

    all_rows = []
    if not master_path or not master_path.exists():
        log.warning("Arquivo 'Vendas You Go 25.xlsx' nao encontrado. Pulando.")
        return all_rows

    log.info(f"Lendo dados historicos You Go 2025 de: {master_path}")
    try:
        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
        ws = wb.active
        count = 0

        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if not row or not row[0]:
                continue

            def g(i):
                v = row[i] if i < len(row) else None
                return "" if v is None else str(v).strip()

            # Data (col 10)
            data_val = row[10] if len(row) > 10 else None
            if isinstance(data_val, (_dt.datetime, _dt.date)):
                data_br = data_val.strftime("%d/%m/%Y")
            else:
                data_br = str(data_val or "").strip()

            # Hora (col 11)
            hora_val = row[11] if len(row) > 11 else None
            if isinstance(hora_val, _dt.time):
                hora_br = hora_val.strftime("%H:%M:%S")
            elif isinstance(data_val, _dt.datetime):
                hora_br = data_val.strftime("%H:%M:%S")
            else:
                hora_br = "12:00:00"

            # Valor pago (col 9)
            total_r = 0.0
            try:
                v = str(row[9] if len(row) > 9 else 0).replace("R$", "").replace(" ", "").strip()
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")
                total_r = float(v)
                if total_r >= 1000 and total_r == int(total_r):
                    total_r = total_r / 100.0
            except Exception:
                total_r = 0.0

            serial        = g(1)
            matricula     = g(3)
            nome_terminal = g(4) or serial
            pagamento     = g(5) or "PRIVATE LABEL"
            produto       = g(6) or "1 Pulso(s)"
            mola_id       = g(7)
            usuario       = g(12)
            nsu_raw       = g(13)
            nsu = nsu_raw if nsu_raw else (f"PB-{serial}|{data_br}|{hora_br}" if serial else "")
            val_str       = str(total_r).replace(".", ",")

            all_rows.append([
                "LAVAÍ - You Go",        # [0]  Cliente
                nome_terminal,           # [1]  Máquina
                "PayBlu (Private Label)",# [2]  Modelo
                "PayBlu",                # [3]  Fabricante
                pagamento,               # [4]  Pagamento
                produto,                 # [5]  Produto
                mola_id,                 # [6]  Mola
                val_str,                 # [7]  Venda R$
                "Nao Informado",         # [8]  Preco R$
                val_str,                 # [9]  Total
                "Nao utilizado",         # [10] Cod Promocional
                data_br,                 # [11] Data
                hora_br,                 # [12] Hora
                matricula,               # [13] N Logico
                nsu,                     # [14] NSU
                "",                      # [15] Autorizacao
                "Private Label",         # [16] Tipo Cartao
                "PayBlu",                # [17] Rede
                "Private Label",         # [18] Bandeira
                usuario,                 # [19] Usuario
                "",                      # [20] No Cartao
                matricula,               # [21] Matricula
            ])
            count += 1

        log.info(f"You Go 2025: {count} transacoes carregadas de '{master_path.name}'.")
    except Exception as e:
        log.error(f"Erro ao ler You Go 2025 '{master_path}': {e}")

    return all_rows


async def coletar_sq_api():
    log.info("Buscando dados de faturamento de moeda e aplicativo do SQInsights...")
    token = None
    
    try:
        async with async_playwright() as p:
            try:
                browser = await p.chromium.connect_over_cdp("http://127.0.0.1:9222")
                target_page = None
                for context in browser.contexts:
                    for page in context.pages:
                        if "sqinsights.com" in page.url:
                            target_page = page
                            break
                    if target_page:
                        break
                if not target_page:
                    log.warning("Aba do SQInsights nao encontrada no Chrome. Moeda/Aplicativo nao serao atualizados.")
                    await browser.close()
                    return []
                    
                ls_str = await target_page.evaluate("() => localStorage.getItem('ember_simple_auth_session')")
                ls_data = json.loads(ls_str)
                token = ls_data.get("authenticated", {}).get("token")
                await browser.close()
            except Exception as e:
                log.warning(f"Erro ao obter token do SQInsights via CDP: {e}")
                return []
    except Exception as e:
        log.warning(f"Erro ao iniciar Playwright para CDP: {e}")
        return []

    if not token:
        log.warning("Token de autenticacao do SQInsights nao encontrado no localStorage.")
        return []

    now = datetime.now(FUSO_SP)
    start_date = now.replace(day=1).strftime("%Y-%m-%d")
    end_date = now.strftime("%Y-%m-%d")
    
    url = f"https://api.sqinsights.com/reports/AUDIT_DETAILED_VENDING/download?&startDate={start_date}&endDate={end_date}&startTime=00:00&endTime=23:59&organizationId=792990&rooms=6512,8082,7663,6821,7662,7288,8440,7927,9137,6822,7664,9138"
    
    req = urllib.request.Request(
        url,
        headers={
            "alliancels-auth-token": token,
            "x-api-key": "4da79517795f579f1717d55b25fb1e9d",
            "alliancels-organization-id": "792990",
            "app": "INSIGHTS",
            "referer": "https://sqinsights.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "accept": "application/json, text/javascript, */*; q=0.01"
        }
    )
    
    try:
        # Request S3 URL
        with urllib.request.urlopen(req, timeout=30) as response:
            res_content = response.read().decode('utf-8').strip()
            s3_url = res_content.strip('"')
            
            # Download actual CSV
            s3_req = urllib.request.Request(
                s3_url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                }
            )
            with urllib.request.urlopen(s3_req, timeout=30) as s3_response:
                csv_bytes = s3_response.read()
                
            csv_content = csv_bytes.decode('utf-8', errors='replace')
            return processar_csv_sq(csv_content, start_date)
            
    except Exception as e:
        log.error(f"Erro ao baixar ou processar relatorio do SQInsights: {e}")
        return []


# ─── Coleta principal ─────────────────────────────────────────────────────────

# Cabeçalho padrão das 22 colunas
CSV_HEADER = "Cliente;Máquina;Modelo;Fabricante;Pagamento;Produtos;Mola;Venda (R$);Preço (R$);Total;Código Promocional;Data;Hora;Nº Logico;NSU;Autorização;Tipo Cartão;Rede;Bandeira;Usuário;Nº Cartão;Matricula"

def salvar_fonte_local(rows, js_path, var_name):
    """Salva rows como .js (window.VAR = CSV) para leitura local pelo dashboard."""
    if not rows:
        log.warning("salvar_fonte_local: sem dados para " + js_path.name)
        return

    cols = []
    cols.append(CSV_HEADER)
    for r in rows:
        padded = list(r) + [""] * (22 - len(r))
        cols.append(";".join(str(c) for c in padded[:22]))

    csv_text = "\n".join(cols)
    escaped  = csv_text.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
    js_out   = "window." + var_name + " = `\n" + escaped + "`;\n"

    try:
        js_path.parent.mkdir(parents=True, exist_ok=True)
        js_path.write_text(js_out, encoding="utf-8")
        log.info("CSV local salvo: " + js_path.name + " (" + str(len(rows)) + " linhas)")
    except Exception as e:
        log.error("Erro ao salvar " + js_path.name + ": " + str(e))

def carregar_fonte_local(caminho):
    if not caminho.exists():
        return []
    try:
        content = caminho.read_text(encoding="utf-8")
        start_idx = content.find("`")
        end_idx = content.rfind("`")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            csv_data = content[start_idx + 1:end_idx]
            reader = csv.reader(csv_data.splitlines(), delimiter=";")
            rows = []
            for r in reader:
                if not r or r[0].strip() == "Cliente":
                    continue
                rows.append(r)
            return rows
    except Exception as e:
        log.error(f"Erro ao carregar dados locais de {caminho.name}: {e}")
    return []

def encontrar_repo_git(path_inicial: Path):
    """Sobe na árvore de diretórios a partir de path_inicial até encontrar um .git."""
    p = path_inicial.resolve()
    for _ in range(8):  # máximo 8 níveis acima
        if (p / ".git").is_dir():
            return p
        if p.parent == p:
            break
        p = p.parent
    return None


def publicar_dados_github():
    import subprocess
    log.info("Iniciando publicação automática no GitHub...")

    script_dir = Path(__file__).parent.resolve()

    # ── Repositório ROOT ──────────────────────────────────────────
    # Detecta o repo git que contém o script (sobe na árvore)
    cwd_root = encontrar_repo_git(script_dir)
    if not cwd_root:
        log.error(f"publicar_dados_github: nenhum repositório git encontrado a partir de {script_dir}. Verifique se 'git init' foi executado na pasta correta.")
        return
    log.info(f"Repositório root detectado: {cwd_root}")

    # Detecta a branch atual do root
    branch_root = "master" # fallback padrão
    try:
        res = subprocess.run(["git", "branch", "--show-current"], cwd=str(cwd_root), capture_output=True, text=True, check=True)
        b_str = res.stdout.strip()
        if b_str:
            branch_root = b_str
    except Exception:
        try:
            res = subprocess.run(["git", "rev-parse", "--abbrev-ref", "HEAD"], cwd=str(cwd_root), capture_output=True, text=True, check=True)
            b_str = res.stdout.strip()
            if b_str:
                branch_root = b_str
        except Exception:
            pass
    log.info(f"Branch atual detectada no root: {branch_root}")

    files_root = [
        "vmpay_local.js",
        "vendtef_local.js",
        "payblu_local.js",
        "sqi_local.js",
        "dados_relatorios.json",
    ]

    for f in files_root:
        f_path = script_dir / f
        if not f_path.exists():
            f_path = cwd_root / f
        if f_path.exists():
            try:
                subprocess.run(
                    ["git", "add", "-f", str(f_path)],
                    cwd=str(cwd_root), check=True, capture_output=True, text=True
                )
            except Exception as e:
                log.error(f"Erro ao adicionar {f} no root: {e}")
        else:
            log.warning(f"Arquivo {f} não encontrado, pulando.")

    now_str = datetime.now(FUSO_SP).strftime("%Y-%m-%d %H:%M:%S")
    try:
        status_res = subprocess.run(
            ["git", "status", "--porcelain"],
            cwd=str(cwd_root), check=True, capture_output=True, text=True
        )
        if status_res.stdout.strip():
            try:
                subprocess.run(["git", "pull", "--rebase", "origin", branch_root],
                               cwd=str(cwd_root), check=True, capture_output=True, text=True)
            except Exception as pe:
                log.warning(f"Aviso ao dar pull --rebase no root ({branch_root}): {pe}")
            subprocess.run(
                ["git", "commit", "-m", f"Auto-update dashboard (root) - {now_str}"],
                cwd=str(cwd_root), check=True, capture_output=True, text=True
            )
            log.info("Commit realizado no root.")
            subprocess.run(["git", "push", "origin", branch_root],
                           cwd=str(cwd_root), check=True, capture_output=True, text=True)
            log.info("Push realizado no root.")
        else:
            log.info("Sem alterações no root.")
    except Exception as e:
        log.error(f"Erro no git commit/push do root ({branch_root}): {e}")

    # ── Repositório KPI (subpasta com .git próprio) ───────────────
    kpi_candidates = [script_dir / "kpi", cwd_root / "kpi"]
    cwd_kpi = None
    for kc in kpi_candidates:
        if kc.is_dir() and (kc / ".git").is_dir():
            cwd_kpi = kc
            break

    if cwd_kpi:
        log.info(f"Repositório kpi detectado: {cwd_kpi}")
        
        # Detecta a branch atual do kpi
        branch_kpi = "main" # fallback padrão
        try:
            res = subprocess.run(["git", "branch", "--show-current"], cwd=str(cwd_kpi), capture_output=True, text=True, check=True)
            b_str = res.stdout.strip()
            if b_str:
                branch_kpi = b_str
        except Exception:
            try:
                res = subprocess.run(["git", "rev-parse", "--abbrev-ref", "HEAD"], cwd=str(cwd_kpi), capture_output=True, text=True, check=True)
                b_str = res.stdout.strip()
                if b_str:
                    branch_kpi = b_str
            except Exception:
                pass
        log.info(f"Branch atual detectada no kpi: {branch_kpi}")

        files_kpi = [
            "vmpay_local.js",
            "vendtef_local.js",
            "payblu_local.js",
            "sqi_local.js",
            "dados_relatorios.json",
        ]

        src_json = script_dir / "dados_relatorios.json"
        dest_json = cwd_kpi / "dados_relatorios.json"
        if src_json.exists():
            try:
                dest_json.write_text(src_json.read_text(encoding="utf-8"), encoding="utf-8")
            except Exception as e:
                log.error(f"Erro ao copiar dados_relatorios.json para kpi: {e}")

        for f in files_kpi:
            f_path = cwd_kpi / f
            if f_path.exists():
                try:
                    subprocess.run(
                        ["git", "add", "-f", f],
                        cwd=str(cwd_kpi), check=True, capture_output=True, text=True
                    )
                except Exception as e:
                    log.error(f"Erro ao adicionar {f} no kpi: {e}")
            else:
                log.warning(f"kpi/{f} não encontrado, pulando.")

        try:
            status_res = subprocess.run(
                ["git", "status", "--porcelain"],
                cwd=str(cwd_kpi), check=True, capture_output=True, text=True
            )
            if status_res.stdout.strip():
                try:
                    subprocess.run(["git", "pull", "--rebase", "origin", branch_kpi],
                                   cwd=str(cwd_kpi), check=True, capture_output=True, text=True)
                except Exception as pe:
                    log.warning(f"Aviso ao dar pull --rebase no kpi ({branch_kpi}): {pe}")
                subprocess.run(
                    ["git", "commit", "-m", f"Auto-update dashboard (kpi) - {now_str}"],
                    cwd=str(cwd_kpi), check=True, capture_output=True, text=True
                )
                log.info("Commit realizado no kpi.")
                subprocess.run(["git", "push", "origin", branch_kpi],
                               cwd=str(cwd_kpi), check=True, capture_output=True, text=True)
                log.info("Push realizado no kpi.")
            else:
                log.info("Sem alterações no kpi.")
        except Exception as e:
            log.error(f"Erro no git commit/push do kpi ({branch_kpi}): {e}")
    else:
        log.info("Pasta kpi/ sem repositório git próprio — pulando push kpi.")


async def coletar_tudo():
    log.info("=" * 60)
    log.info("Iniciando coleta de relatorios (Mensal)")
    log.info("=" * 60)

    rows = []
    payblu_rows = []  # inicializa caso o bloco try falhe antes de coletar_payblu
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            locale="pt-BR",
            timezone_id="America/Sao_Paulo",
        )
        page = await context.new_page()

        try:
            log.info("Acessando login do ERP Vending...")
            await page.goto(URL_ERP, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
            await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
            
            ok = await fazer_login(page)
            if not ok:
                log.error("Falha ao logar no ERP Vending após 3 tentativas. VendTEF e PayBlu não serão coletados nesta execução.")
                # Não encerra — continua para API VMPay, SQI e VendPago
                raise Exception("ERP login failed — skip VendTEF/PayBlu")

            # Estabelecer SSO com VendTEF
            log.info("Estabelecendo sessao no VendTEF via link SSO...")
            await page.wait_for_selector("a#navbarVendtef", timeout=10000)
            payblu_link = page.locator("a#navbarVendtef")
            href = await payblu_link.get_attribute("href")
            if href:
                await page.goto(href, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
                await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
                await asyncio.sleep(3)
            else:
                log.warning("Navbar link do VendTEF nao contem href. Tentando URL direta...")
                await page.goto(URL_DOWNLOAD, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
                await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)

            # Acessar relatorioVendasGeralDownload
            log.info(f"Acessando pagina de download: {URL_DOWNLOAD}")
            await page.goto(URL_DOWNLOAD, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
            await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
            await asyncio.sleep(2)
            
            await dispensar_modal(page)
            
            # Clicar em Continuar para gerar o relatorio do mês corrente
            log.info("Clicando em 'Continuar'...")
            await page.locator("input[value='Continuar']").click(timeout=10000)
            await page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
            await asyncio.sleep(4)
            
            # Clicar em Download
            log.info("Clicando no botao 'Download'...")
            temp_zip = Path(__file__).parent / "temp_vendas.zip"
            btn_locator = page.locator("a:has-text('Download'), button:has-text('Download'), input[value='Download']").first
            async with page.expect_download(timeout=20000) as download_info:
                await btn_locator.click()
                
            download = await download_info.value
            await download.save_as(str(temp_zip))
            log.info(f"Relatorio ZIP baixado com sucesso.")

            # Extrair e processar CSV
            with zipfile.ZipFile(temp_zip, 'r') as zip_ref:
                csv_name = zip_ref.namelist()[0]
                content = zip_ref.read(csv_name).decode('latin-1')
                
            reader = csv.reader(content.splitlines(), delimiter=';')
            for i, r in enumerate(reader):
                if i < 6:
                    continue
                if len(r) >= 22:
                    rows.append([cell.strip() for cell in r])
                    
            log.info(f"Total de {len(rows)} transacoes encontradas no relatorio mensal.")

            if temp_zip.exists():
                temp_zip.unlink()

            # Coletar PayBlu com a mesma sessão autenticada
            payblu_rows = await coletar_payblu(page)

        except Exception as e:
            log.error(f"Erro durante extracao ERP/VendTEF/PayBlu: {e}")
            log.warning("Continuando pipeline com as demais fontes (API VMPay, SQInsights, VendPago).")
            payblu_rows = []
            # Limpa zip temporário se existir
            _tmp = Path(__file__).parent / "temp_vendas.zip"
            if _tmp.exists():
                try: _tmp.unlink()
                except: pass
        finally:
            await browser.close()

    # Coletar dados da API VMPay Cashless, SQInsights, VendPago Excel e unificar com os dados raspados
    api_rows = coletar_vmpay_api()
    # excel_rows removido: VMPay local não é mais usado; fonte única = API VMPay
    sq_rows = await coletar_sq_excel()
    vendpago_excel_rows = coletar_vendpago_excel()
    yougo_excel_rows = coletar_yougo_excel()  # histórico PayBlu You Go 2025

        # Carrega dados VMPay já existentes para mesclar (evita perder histórico se a API falhar ou se trouxer período parcial)
    vmpay_antigos = []
    for p_cache in [CSV_VMPAY_LOCAL, Path(__file__).parent / "kpi" / "vmpay_local.js", Path(__file__).parent.parent / "vmpay_local.js"]:
        if p_cache.exists() and p_cache.stat().st_size >= 100:
            vmpay_antigos = carregar_fonte_local(p_cache)
            if vmpay_antigos:
                break

    # Deduplicar cada fonte antes de enviar (evitar duplicatas dentro do mesmo lote)
    api_rows_to_merge = api_rows if api_rows is not None else []
    portal_rows_dedup   = merge_and_deduplicate(rows + vendpago_excel_rows, [])
    
    # Se a API teve sucesso, mescla o novo lote com o histórico. Se falhou, preserva o histórico intacto.
    if api_rows is not None:
        vmpay_rows_dedup = merge_and_deduplicate(vmpay_antigos + api_rows_to_merge, [])
        log.info(f"VMPay: mesclados {len(api_rows_to_merge)} novos registros da API com {len(vmpay_antigos)} históricos (Total: {len(vmpay_rows_dedup)}).")
    else:
        vmpay_rows_dedup = merge_and_deduplicate(vmpay_antigos, [])
        log.warning(f"VMPay: mantendo {len(vmpay_rows_dedup)} registros históricos devido a falha na API.")
        
    sq_rows_dedup       = merge_and_deduplicate(sq_rows, [])
    # PayBlu: deduplica histórico Excel junto com dados do portal (mesma fonte, evita duplicatas)
    payblu_rows_dedup   = merge_and_deduplicate(payblu_rows + yougo_excel_rows, [])

    total_count = (len(portal_rows_dedup) + len(vmpay_rows_dedup) + len(sq_rows_dedup)
                   + len(payblu_rows_dedup))

        # Determinar diretórios de gravação baseados no repositório git
    root_dir = encontrar_repo_git(Path(__file__).parent)
    if not root_dir:
        root_dir = Path(__file__).parent
    kpi_dir = root_dir / "kpi"

    # Salva JSON de status para manter compatibilidade
    payload = {
        "gerado_em": datetime.now(FUSO_SP).isoformat(),
        "proxima_coleta": None,
        "status": "ok" if total_count else "erro",
        "total_transacoes": total_count,
        "portal_transacoes": len(portal_rows_dedup),
        "api_transacoes": len(api_rows_to_merge),
        "excel_transacoes": len(vendpago_excel_rows) + len(yougo_excel_rows),  # excel VMPay removido
        "sq_transacoes": len(sq_rows_dedup),
        "payblu_transacoes": len(payblu_rows_dedup)
    }
    
    (root_dir / "dados_relatorios.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if kpi_dir.is_dir() and kpi_dir != root_dir:
        (kpi_dir / "dados_relatorios.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── Todas as fontes salvas como JS local — zero envio para o Sheets ──────
    salvar_fonte_local(portal_rows_dedup,   root_dir / "vendtef_local.js",  "LAVAI_VENDTEF_DATA")
    salvar_fonte_local(vmpay_rows_dedup,    root_dir / "vmpay_local.js",    "LAVAI_VMPAY_DATA")
    salvar_fonte_local(payblu_rows_dedup,   root_dir / "payblu_local.js",   "LAVAI_PAYBLU_DATA")
    salvar_fonte_local(sq_rows_dedup,       root_dir / "sqi_local.js",      "LAVAI_SQI_DATA")

    # Também salvar na subpasta 'kpi' se ela existir (para manter a branch main de deploy atualizada)
    if kpi_dir.is_dir() and kpi_dir != root_dir:
        salvar_fonte_local(portal_rows_dedup,   kpi_dir / "vendtef_local.js",  "LAVAI_VENDTEF_DATA")
        salvar_fonte_local(vmpay_rows_dedup,    kpi_dir / "vmpay_local.js",    "LAVAI_VMPAY_DATA")
        salvar_fonte_local(payblu_rows_dedup,   kpi_dir / "payblu_local.js",   "LAVAI_PAYBLU_DATA")
        salvar_fonte_local(sq_rows_dedup,       kpi_dir / "sqi_local.js",      "LAVAI_SQI_DATA")
        log.info("Arquivos locais da pasta 'kpi' também foram atualizados automaticamente.")

    log.info("Todos os arquivos JS locais atualizados. Nenhum envio para o Google Sheets.")
    
    # Publicar dados automaticamente no GitHub
    try:
        publicar_dados_github()
    except Exception as e:
        log.error(f"Erro ao publicar dados no GitHub: {e}")
    
    return payload


# ─── Scheduler ────────────────────────────────────────────────────────────────

def job():
    asyncio.run(coletar_tudo())


if __name__ == "__main__":
    modo_agora = "--agora" in sys.argv

    if modo_agora:
        asyncio.run(coletar_tudo())
        log.info("Modo --agora: coleta concluída, encerrando.")
        sys.exit(0)
    else:
        log.info("Modo continuo: coleta agendada a cada 3 minutos.")
        log.info("Pressione Ctrl+C para parar.\n")

        job()

        schedule.every(3).minutes.do(job)
        while True:
            schedule.run_pending()
            time.sleep(30)
